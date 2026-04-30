"""Build stored outline/hint JSON for workspace template assets (DOCX, PDF, XLSX)."""

from __future__ import annotations

import json
import re
from pathlib import Path

from app.ingestion.docx_extract import extract_docx_text_and_outline
from app.ingestion.pdf_extract import extract_pdf_text


_NUM_HEADING = re.compile(r"^(\d+(\.\d+)*\.?\s+|[A-Z]\.\s+|[IVXLCDM]+\.\s+)\S.+")
_SECTION_PREFIX = re.compile(
    r"^(SECTION|PART|ANNEX|APPENDIX|ATTACHMENT|EXHIBIT)\s+[A-Z0-9]+\b.*",
    re.I,
)


def _guess_section_lines_from_text(text: str, max_lines: int = 35) -> list[str]:
    """Heuristic headings for non-Word sources (PDF plain text has no style metadata)."""
    seen: set[str] = set()
    out: list[str] = []
    for raw in text.splitlines():
        s = raw.strip()
        if len(s) < 4 or len(s) > 120:
            continue
        low = s.lower()
        if low.startswith(("page ", "copyright", "classified")):
            continue
        if re.match(r"^page\s+\d+(\s+of\s+\d+)?\b", low):
            continue

        matched = False
        if _NUM_HEADING.match(s):
            matched = True
        elif _SECTION_PREFIX.match(s):
            matched = True
        elif s.isupper() and " " in s and sum(c.isdigit() for c in s) < len(s) * 0.25:
            matched = True

        if matched and s not in seen:
            seen.add(s)
            out.append(s)
    return out[:max_lines]


def _xlsx_outline(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        title = sheet.title or "Sheet1"
        rows: list[list[str]] = []
        for i, row in enumerate(sheet.iter_rows(max_row=120, max_col=32, values_only=True)):
            if i > 100:
                break
            cells: list[str] = []
            for c in row:
                if c is None:
                    cells.append("")
                else:
                    cells.append(str(c).strip()[:500])
            if any(x for x in cells):
                rows.append(cells)
    finally:
        wb.close()

    headers = rows[0] if rows else []
    non_empty_headers = [h for h in headers if h]
    return {
        "kind": "xlsx",
        "sheet": title,
        "headers": non_empty_headers[:32],
        "sample_rows": rows[:80],
    }


def build_template_outline_json(path: Path, filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".docx"):
        text, headings = extract_docx_text_and_outline(path)
        return json.dumps(
            {
                "kind": "docx",
                "headings": headings,
                "text_excerpt": (text or "")[:10000],
            }
        )
    if lower.endswith(".pdf"):
        text, meta = extract_pdf_text(path)
        guessed = _guess_section_lines_from_text(text or "")
        return json.dumps(
            {
                "kind": "pdf",
                "headings": guessed,
                "text_excerpt": (text or "")[:12000],
                "pages": meta.get("pages"),
            }
        )
    if lower.endswith(".xlsx"):
        payload = _xlsx_outline(path)
        return json.dumps(payload)
    raise ValueError(f"Unsupported template type: {filename}")
